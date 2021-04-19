from datetime import datetime
import pandas as pd
import glob
import openpyxl
from static import *

# paths
ASSETS_PATH  = paths['ASSETS_PATH']

def master_map():
    path = glob.glob(os.path.join(ASSETS_PATH, 'product_mapping.xlsx'))[0]
    df_map = pd.read_excel(path, sheet_name='map')
    df = df_map[['Source', 'Amex_Parent', 'Product', 'UM_Segment_Split', 'UM_Message', 'UM_CardProduct']]
    return df

def check_product_mapping(df):
    unmapped = df[(df['UM_Segment_Split'].isin(['<NA>'])) |
                            (df['UM_Message'].isin(['<NA>'])) |
                            (df['UM_CardProduct'].isin(['<NA>']))]

    unmapped_unique = unmapped[['Product', 'UM_Segment_Split', 'UM_Message', 'UM_CardProduct']].drop_duplicates()

    total_len = len(unmapped)
    unique_len = len(unmapped_unique)

    if total_len>0:
        print(f"Missing Product Mapping")
        print(f"\tTotal Rows Unmapped: {total_len}")
        print(f"\tUnique Rows Unmapped: {unique_len}\n")

    return None

def check_media_type_mapping(df):
    source = df['Source'].unique()[0]
    unmapped = df[df['UM_Media_Type'].isin(['<NA>'])]
    total_len = len(unmapped)

    if total_len>0:
        if source.lower()=='kantar':
            unmapped_unique = unmapped[['MEDIA', 'MEDIA_TYPE', 'UM_Media_Type']].drop_duplicates()
            print(f"Missing UM Media Type")
            print(f"\t{unmapped_unique}\n")

        elif source.lower()=='pathmatics':
            unmapped_unique = unmapped[['Device', 'UM_Media_Type']].drop_duplicates()
            print(f"Missing UM Media Type")
            print(f"\t{unmapped_unique}\n")
    return None

def check_amex_parent_mapping(df):
    source = df['Source'].unique()[0]
    unmapped = df[df['Amex_Parent'].isin(['<NA>'])]
    total_len = len(unmapped)

    if total_len>0:
        if source.lower()=='kantar':
            unmapped_unique = unmapped[['AMEX_PARENT', 'Amex_Parent']].drop_duplicates()
            print(f"Missing Amex Parent")
            print(f"\t{unmapped_unique}\n")

        elif source.lower()=='pathmatics':
            unmapped_unique = unmapped[['Advertiser', 'Amex_Parent']].drop_duplicates()
            print(f"Missing Amex Parent")
            print(f"\t{unmapped_unique}\n")

    return None

def export_missing_products(df):
    unmapped = df.copy()
    unmapped = unmapped[['report_date', 'Source', 'Amex_Parent', 'Product', 'UM_Segment_Split', 'UM_Message', 'UM_CardProduct']].drop_duplicates()

    unmapped = unmapped[(unmapped['UM_Segment_Split'].isin(['<NA>'])) |
                        (unmapped['UM_Message'].isin(['<NA>'])) |
                        (unmapped['UM_CardProduct'].isin(['<NA>']))]

    unmapped.rename(columns={'report_date':'date_added'}, inplace=True)

    len_ = len(unmapped)

    if len_>0:
        print("Exporting Missing Products")
        print(f"\tLength: {len_}")
        # unmapped.to_excel(os.path.join(ASSETS_PATH, 'MISSING_product_mapping.xlsx'), index=False)

        path = glob.glob(os.path.join(paths['ASSETS_PATH'], 'product*'))[0]
        wb = openpyxl.load_workbook(path)
        ws = wb['map']
        starting_row = ws.max_row + 1
        no_of_cols = [c for c in range(ws.max_column)]

        for i in range(len_):
            for c in no_of_cols:
                ws.cell(starting_row + i, c+1).value = unmapped.iloc[i, c]

        wb.save(path)
        wb.close()

    else:
        print("All Products Mapped")

    return None


def special_case_adjustments(df):
    # Citigroup
    df['UM_CardProduct'] = df.apply(lambda x: 'Card/Product' if (x['Amex_Parent'] == 'Citigroup Inc' and x['Product'] == 'Apple Pay Apple Pay') else x['UM_CardProduct'] , axis=1)
    df['UM_Message'] = df.apply(lambda x: 'rewards plus credit cards' if (x['Amex_Parent'] == 'Citigroup Inc' and x['Product'] == 'Apple Pay Apple Pay') else x['UM_Message'] , axis=1)
    df['UM_Segment_Split'] = df.apply(lambda x: 'B2C' if (x['Amex_Parent'] == 'Citigroup Inc' and x['Product'] == 'Apple Pay Apple Pay') else x['UM_Segment_Split'] , axis=1)
    df['Product'] = df.apply(lambda x: 'Citi Rewards Plus' if (x['Amex_Parent'] == 'Citigroup Inc' and x['Product'] == 'Apple Pay Apple Pay') else x['Product'] , axis=1)

    # Mastercard to PayPal
    df['Amex_Parent'] = df.apply(lambda x: 'PayPal Holdings Inc' if (x['Amex_Parent'] == 'Mastercard Intl Inc' and x['Product'] == 'PayPal Cash : Debit Card : MasterCard') else x['Amex_Parent'], axis=1)

    # Marke REMOVE for Kabbage < 10/1/2020
    cols = ['UM_Segment_Split', 'UM_Message', 'UM_CardProduct']
    for c in cols:
        df[c] = df.apply(lambda x: 'REMOVE' if 'kabbage' in x['Product'].lower() and x['Month'] < datetime(2020,10,1) else x[c], axis=1)

    return df
