from openpyxl.styles.borders import Border, Side
from openpyxl import styles, formatting
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import numpy as np
import openpyxl
import glob
import os

percent_format = '0.0%'
number_format = '#,###0'

thin_border = Border(right=Side(style='thin'))
all_borders = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

red_color = '00FF0000'
red_fill = styles.PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')

light_red_color = 'FF7676'
light_red_fill = styles.PatternFill(start_color=light_red_color, end_color=light_red_color, fill_type='solid')

ws_args = {
    'search':{
        'advertiser': {
            'insert_row': 1,
            'start_parse': 2,
            'impr_col': 9
                        },
        'campaign': {
            'insert_row': 1,
            'start_parse': 3,
            'impr_col': 10
                    },
        'week': {
            'insert_row': 1,
            'start_parse': 4,
            'impr_col': 11
                }
    },
    'social':{
        'campaign':{
            'insert_row': 1,
            'start_parse': 2,
            'impr_col': 8
                    },
        'adset':{
            'insert_row': 1,
            'start_parse': 3,
            'impr_col': 9
                    },
        'week':{
            'insert_row': 1,
            'start_parse': 4,
            'impr_col': 10
                    },
        'tab_color_hex': {
            'fb': '00B2FF',
            'li': '2867B2',
            'pi': 'E60023',
            'tw': '1DA1F2'
                    }
    },
    'digital':{
        'campaign': {
            'insert_row': 1,
            'start_parse': 4,
            'impr_col': 7
                        },
        'placement': {
            'insert_row': 1,
            'start_parse': 5,
            'impr_col': 8
                    },
        'week': {
            'insert_row': 1,
            'start_parse': 5,
            'impr_col': 8
                }
    },
          }


def tab_color(media, wb_obj):
    if media == 'social':
        for sht in wb_obj.sheetnames:
            sht_plat = sht.split('_')[0]
            if sht_plat in ws_args[media]['tab_color_hex'].keys():
                ws = wb_obj[sht]
                ws.sheet_properties.tabColor = ws_args[media]['tab_color_hex'][sht_plat]
    return wb_obj


def format_qa(media, qa_filepath):
    # create sb obj
    wb = openpyxl.load_workbook(qa_filepath)

    # main loop
    for sht in wb.sheetnames:
        try:
            sht_parse = sht.split('_')[-1]

            # insert row for sheets in ws_args
            ws = wb[sht]
            ws.insert_rows(ws_args[media][sht_parse]['insert_row'])

            # add the metric name to the first row
            for c in range(ws_args[media][sht_parse]['start_parse'], ws.max_column, 3):
                val = ws.cell(2, c).value
                metric = val.split('_')[0]
                ws.cell(1, c).value = metric
                ws.merge_cells(start_row=1, start_column=c, end_row=1, end_column=c+2) # merge cells with metric name
                ws.cell(1, c).alignment = Alignment(horizontal='center')

                # add border
                for r in range(1, ws.max_row + 1):
                    # add a border every 2 cells from the start_parse
                    ws.cell(r, c+2).border = thin_border

                    # tag threshold
                    perc_diff = ws.cell(r, c+2).value
                    impr_val = ws.cell(r, ws_args[media][sht_parse]['impr_col']).value
                    try:
                        # apply formatting
                        ws.cell(r, c).number_format = number_format
                        ws.cell(r, c+1).number_format = number_format
                        ws.cell(r, c+2).number_format = percent_format

                        # matterkind exception
                        site_served_col = 1
                        site_name_col = ws_args[media][sht_parse]['start_parse'] - 1
                        metric_col = [site_name_col + 4, site_name_col + 5, site_name_col + 6]

                        # rs > ui
                        if ws.cell(r, site_name_col).value == 'Matterkind' and c in metric_col:
                            continue
                        else:
                            if int(ws.cell(r, c).value) > int(ws.cell(r, c+1).value):
                                ws.cell(r, c).fill = light_red_fill

                        # % diff > +/- 2% for impr > 100
                        if ws.cell(r, site_name_col).value == 'Matterkind' and c in metric_col:
                            continue
                        else:
                            if np.abs(perc_diff) >= 0.02:
                                if ws.cell(r, site_served_col).value:
                                    ws.cell(r, c+2).fill = red_fill
                                elif float(impr_val) > 100.0:
                                    ws.cell(r, c+2).fill = red_fill
                    except TypeError:
                        continue
                    except ValueError:
                        continue

            # change column name
            for c in range(ws_args[media][sht_parse]['start_parse'], ws.max_column + 1):
                val = ws.cell(2, c).value
                source = val.split('_')[-1]
                ws.cell(2, c).value = {True:'% diff', False: source}[source == 'diff']
                ws.cell(2, c).border = all_borders

            # freeze pane
            ws.freeze_panes = f"{get_column_letter(ws_args[media][sht_parse]['start_parse'])}3"

        except KeyError:
#             print('error')
            continue

    wb = tab_color(media, wb)
    wb.save(qa_filepath)
    wb.close()
