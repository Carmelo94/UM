from datetime import datetime, timedelta
import pandas as pd
import glob
import os
import time
import openpyxl

import sys
sys.path.append('../utils')
from SharePoint import *
from hlpr import *
from static import *

def metric_update():
    '''Download metric mapping from SharePoint and append new metrics from RedShift database.'''

    start = time.time()

    # connect to SharePoint
    app = SharePoint('AmericanExpressUSGABM')
    app.list_contents(SHAREPOINT_MAPPING_PATH)
    app.download_files(ASSETS_PATH)

    # get metrics from database
    df_rs_metrics = run_qry(REDSHIFT_METRICS_QUERY)
    # df_rs_metrics.to_excel(os.path.join(ASSETS_PATH, 'metrics.xlsx'), index=False)
    df_rs_metrics.dropna(inplace=True) # drop null results

    # import the SharePoint metric file
    pth = glob.glob(os.path.join(ASSETS_PATH, f"*{REDSHIFT_METRICS_FILENAME}*"))[0]
    df_cur_metrics = pd.read_excel(pth, sheet_name=REDSHIFT_METRICS_SHEETNAME)

    # get list of metrics
    metric_field = 'universal_metric' if media=='social' else 'metric'
    rs_metrics = list(df_rs_metrics[metric_field])
    cur_metrics = list(df_cur_metrics[metric_field])

    new_metrics = [m for m in rs_metrics if not(m in cur_metrics)]

    # if there is missing data, create a df
    if len(new_metrics) > 0:
        print(f"new metrics identified: {len(new_metrics)}")

        # initiate openpyxl
        wb = openpyxl.load_workbook(pth)
        ws = wb[REDSHIFT_METRICS_SHEETNAME]

        starting_row = ws.max_row + 1

        if media in ['search', 'digital']:
            for i in range(len(new_metrics)):
                ws.cell(starting_row + i, 1).value = int(SharePoint.dt)
                ws.cell(starting_row + i, 2).value = new_metrics[i]
        elif media=='social':
            rs_metrics_full = df_rs_metrics[df_rs_metrics['universal_metric'].isin(new_metrics)]
            for i in range(len(new_metrics)):
                ws.cell(starting_row + i, 1).value = int(SharePoint.dt)
                ws.cell(starting_row + i, 2).value = rs_metrics_full.iloc[i, 0]
                ws.cell(starting_row + i, 3).value = rs_metrics_full.iloc[i, 1]
                ws.cell(starting_row + i, 4).value = rs_metrics_full.iloc[i, 2]

        wb.save(pth)
        wb.close()

        # app.upload_files(SHAREPOINT_MAPPING_PATH, pth)
    else:
        print('no metrics to update')

    print(f"\nruntime: {np.round(time.time() - start, 2)} seconds")

if __name__ == '__main__':
    medias = ['digital', 'search', 'social']
    for media in medias:
        print(f"\nupdating {media} metrics\n{'='*50}")
        SHAREPOINT_DATA_PATH = media_args[media]['sharepoint']['data']
        SHAREPOINT_MAPPING_PATH = media_args[media]['sharepoint']['mapping']
        SHAREPOINT_QA_PATH = media_args[media]['sharepoint']['qa']
        SHAREPOINT_FLAT_PATH = media_args[media]['sharepoint']['flat']
        REDSHIFT_METRICS_QUERY = media_args[media]['redshift']['all_metrics']
        REDSHIFT_METRICS_FILENAME = media_args[media]['redshift']['metric_filename']
        REDSHIFT_METRICS_SHEETNAME = media_args[media]['redshift']['metric_sheetname']
        REDSHIFT_QA_QUERY = media_args[media]['redshift']['qa_qry']
        metric_update()
